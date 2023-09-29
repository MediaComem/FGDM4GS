# -*- coding: utf-8 -*-

"""
geobasisdaten_api.py
Script to write the data from the geobasisdaten API to an Excel file  
Author: [Maxime Collombin]
Date: [29/09/2023]
"""

import requests
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import Cell
from openpyxl.styles.borders import Border, Side

# Make API call and store the response

response = requests.get('https://geobasisdaten.ch/api/v1/data/')

# Store the response in a dataframe

df = pd.DataFrame(columns=['autorité','titre', 'id', 'Modèle minimal', 'Documentation', 'view'])

for item in response.json():
    title = item['titleFr']
    for executiveAuthority in item['executiveAuthorities']:
        authority = executiveAuthority['abbrFr']
    for technical_entry in item['technicalEntries']:
        label = technical_entry['label']
        data_model_url = technical_entry['dataModelUrlFr']
        model_doc_url = technical_entry['modelDocumentationUrlFr']
        has_view = 't' if 'VIEW' in str(data_model_url).upper() else 'f'
        df = df.append({'autorité':authority, 'titre': title, 'id': label, 'Modèle minimal': data_model_url, 'Documentation': model_doc_url, 'view': has_view}, ignore_index=True)

# Save the dataframe as a xlsx file

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Data'
worksheet.append(df.columns.tolist()) # add header row
for index, row in df.iterrows():
    worksheet.append(row.tolist())

# Add an hyperlink to the "Modèle minimal" column

for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "Documentation" column

for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')
            
# Adjust column width and wrap text
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    if adjusted_width > 100:
        adjusted_width = 100
    worksheet.column_dimensions[column].width = adjusted_width
    worksheet.column_dimensions[column].auto_size = True
    for cell in col:
        cell.alignment = cell.alignment.copy(wrap_text=True)

table = Table(displayName="Table1", ref=f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}")
style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
worksheet.add_table(table)

# Save the workbook

workbook.save('geobasisdaten.xlsx')
