# -*- coding: utf-8 -*-

"""
geobasisdaten_api.py
Script to write the data from the geobasisdaten API to an Excel file  
Author: [Maxime Collombin]
Date: [29/09/2023]
"""

import requests
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from owslib.csw import CatalogueServiceWeb
from xml.etree import ElementTree as ET
from urllib.parse import urlparse, parse_qs, urlunparse, urlencode

# Function to find and print element
def find_and_print(element, path):
    found_element = element.find(path, ns)
    return found_element

# Make API call and store the response
response = requests.get('https://geobasisdaten.ch/api/v1/data/')

# Store the response in a dataframe
df = pd.DataFrame(columns=['autorité','titre', 'id', 'Modèle minimal', 'Lien geocat', 'CSW Request', 'Documentation', 'view', 'GetStyles', 'GetFeatureInfo'])

# Initialize the counters

total_entries_count = 0
data_model_url_count = 0

for item in response.json():
    title = item['titleFr']
    for executiveAuthority in item['executiveAuthorities']:
        authority = executiveAuthority['abbrFr']
    for technical_entry in item['technicalEntries']:

        # Increment the total entries counter
        total_entries_count += 1

        label = technical_entry['label']
        data_model_url = technical_entry['dataModelUrlFr']

        # Increment the counter if data_model_url is not null and not empty
        if data_model_url:
            data_model_url_count += 1

        model_doc_url = technical_entry['modelDocumentationUrlFr']
        geocat_links = [metadata['urlFr'] for metadata in technical_entry['metadataUrls'] if metadata['urlFr'] is not None]  # Ensure urlFr is not None
        has_view = 't' if 'VIEW' in str(data_model_url).upper() else 'f'
        for geocat_link in geocat_links:  # Loop through the geocat links
            last_part = geocat_link.split('/')[-1]  # Get the last part of the geocat link
            csw_request = f"https://www.geocat.ch/geonetwork/srv/fre/csw?service=CSW&version=2.0.2&request=GetRecordById&id={last_part}&outputSchema=http://www.geocat.ch/2008/che&elementSetName=full"

            # CSW to WMS
            csw = CatalogueServiceWeb(csw_request)
            
            csw.getrecordbyid(id=[last_part], esn='full', outputschema='http://www.geocat.ch/2008/che', format='application/xml')

            root = ET.fromstring(csw.response)
            ns = {'gmd': 'http://www.isotc211.org/2005/gmd', 'gco': 'http://www.isotc211.org/2005/gco', 'xsi': 'http://www.w3.org/2001/XMLSchema-instance', 'che': 'http://www.someurl.org'}

            identification_info = find_and_print(root, ".//gmd:identificationInfo")
            identifier = find_and_print(identification_info, ".//gmd:identifier")
            code = find_and_print(identifier, ".//gmd:code/gco:CharacterString")

            # Check if code is not None and code.text is not None before generating URLs
            if code is not None and code.text is not None:
                distribution_info = find_and_print(root, ".//gmd:distributionInfo")
                urls = distribution_info.findall(".//gmd:URL", ns)

                getstyles_url = None
                getfeatureinfo_url = None

                if urls:
                    for url in urls:
                        if url.text and ('WMS' in url.text or 'WFS' in url.text):
                            # Parse the URL and its parameters
                            parsed_url = urlparse(url.text)
                            params = parse_qs(parsed_url.query)

                            # Remove the language parameter
                            params.pop('lang', None)

                            # Replace 'GetCapabilities' with 'GetStyles'
                            params['REQUEST'] = ['GetStyles']

                            # Add 'LAYERS' parameter
                            if code is not None and code.text is not None:
                                params['LAYERS'] = [code.text]

                            # Reconstruct the URL
                            new_query = urlencode(params, doseq=True)
                            getstyles_url = urlunparse(parsed_url._replace(query=new_query))

                            # Replace 'GetStyles' with 'GetFeatureInfo' and add required parameters
                            if code is not None and code.text is not None:
                                params['REQUEST'] = ['GetFeatureInfo']
                                params['QUERY_LAYERS'] = [code.text]
                                params['CRS'] = ['EPSG:2056']
                                params['BBOX'] = ['2531423.89,1155079.22,2532223.89,1155679.22']
                                params['FEATURE_COUNT'] = ['1']
                                params['HEIGHT'] = ['2']
                                params['WIDTH'] = ['2']
                                params['INFO_FORMAT'] = ['text/plain']
                                params['I'] = ['1']
                                params['J'] = ['1']

                            # Reconstruct the URL
                            new_query = urlencode(params, doseq=True)
                            getfeatureinfo_url = urlunparse(parsed_url._replace(query=new_query))

                            # Add the URLs to the dataframe
                            df = df.append({'autorité': authority, 'titre': title, 'id': label, 'Modèle minimal': data_model_url, 'Lien geocat': geocat_link, 'GetRecordById': csw_request, 'Documentation': model_doc_url, 'view': has_view, 'GetStyles': getstyles_url, 'GetFeatureInfo': getfeatureinfo_url}, ignore_index=True)
                            # filter the dataframe to keep only the rows with a non-empty "Modèle minimal" column
                            df = df[df['Modèle minimal'].notna()]

# Save the dataframe as a xlsx file

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Data'
worksheet.append(df.columns.tolist()) # add header row
for index, row in df.iterrows():
    worksheet.append(row.tolist())

# Add an hyperlink to the "Modèle minimal" column

for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "Lien geocat" column

for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "GetRecordById" column
            
for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "Documentation" column

for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')


# Add hyperlink to "GetStyles" column
            
for row in worksheet.iter_rows(min_row=2, min_col=8, max_col=8):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "GetFeatureInfo" column
            
for row in worksheet.iter_rows(min_row=2, min_col=9, max_col=9):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')
            
# Adjust column width and wrap text

for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    if adjusted_width > 100:
        adjusted_width = 100
    worksheet.column_dimensions[column].width = adjusted_width
    worksheet.column_dimensions[column].auto_size = True
    for cell in col:
        cell.alignment = cell.alignment.copy(wrap_text=True)

table = Table(displayName="Table1", ref=f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}")
style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
worksheet.add_table(table)

# Save the workbook

workbook.save('geobasisdaten.xlsx')