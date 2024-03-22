# -*- coding: utf-8 -*-

"""
geobasisdaten_api.py
Script to write the data from the geobasisdaten API to an Excel file  
Author: [Maxime Collombin]
Date: [29/09/2023]
"""

import requests
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import Cell
from openpyxl.styles.borders import Border, Side

# Make API call and store the response

response = requests.get('https://geobasisdaten.ch/api/v1/data/')

# get all the parameters of the API

# print(response.json())

# Store the response in a dataframe

df = pd.DataFrame(columns=['autorité','titre', 'id', 'Modèle minimal', 'Lien geocat', 'CSW Request', 'Documentation', 'view'])

for item in response.json():
    title = item['titleFr']
    for executiveAuthority in item['executiveAuthorities']:
        authority = executiveAuthority['abbrFr']
    for technical_entry in item['technicalEntries']:
        print(len(technical_entry))
        label = technical_entry['label']
        data_model_url = technical_entry['dataModelUrlFr']
        model_doc_url = technical_entry['modelDocumentationUrlFr']
        geocat_links = [metadata['urlFr'] for metadata in technical_entry['metadataUrls'] if metadata['urlFr'] is not None]  # Ensure urlFr is not None
        has_view = 't' if 'VIEW' in str(data_model_url).upper() else 'f'
        for geocat_link in geocat_links:  # Loop through the geocat links
            last_part = geocat_link.split('/')[-1]  # Get the last part of the geocat link
            csw_request = f"https://www.geocat.ch/geonetwork/srv/fre/csw?service=CSW&version=2.0.2&request=GetRecordById&id={last_part}&outputSchema=http://www.geocat.ch/2008/che&elementSetName=full"
            df = df.append({'autorité':authority, 'titre': title, 'id': label, 'Modèle minimal': data_model_url, 'Lien geocat': geocat_link, 'CSW Request': csw_request, 'Documentation': model_doc_url, 'view': has_view}, ignore_index=True)

# Save the dataframe as a xlsx file

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Data'
worksheet.append(df.columns.tolist()) # add header row
for index, row in df.iterrows():
    worksheet.append(row.tolist())

# Add an hyperlink to the "Modèle minimal" column

for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "Lien geocat" column

for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "CSW Request" column
            
for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')

# Add hyperlink to "Documentation" column

for row in worksheet.iter_rows(min_row=2, min_col=7, max_col=7):
    for cell in row:
        if cell.value:
            cell.hyperlink = cell.value
            cell.value = "Lien"
            cell.font = Font(underline='single', color='0000FF')
            
# Adjust column width and wrap text
            
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    if adjusted_width > 100:
        adjusted_width = 100
    worksheet.column_dimensions[column].width = adjusted_width
    worksheet.column_dimensions[column].auto_size = True
    for cell in col:
        cell.alignment = cell.alignment.copy(wrap_text=True)

table = Table(displayName="Table1", ref=f"A1:{get_column_letter(df.shape[1])}{df.shape[0]+1}")
style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
worksheet.add_table(table)

# Save the workbook

workbook.save('test.xlsx')
